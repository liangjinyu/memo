#!/usr/bin/env python 
# -*- coding: utf-8 -*-

from openpyxl.reader.excel import load_workbook
import openpyxl,json

wb = load_workbook(filename=r'test.xlsx')

# 显示有多少张表
print "Worksheet range(s):", wb.get_named_ranges()
print "Worksheet name(s):", wb.get_sheet_names()

sheetnames = wb.get_sheet_names()
ws = wb.get_sheet_by_name(sheetnames[0])
cell_11 = ws["A1"]
print("cess_11 = "+cell_11.value) 
# 显示表名，表行数，表列数
print "Work Sheet Titile:", ws.title
print "Work Sheet Rows:", ws.max_row
print "Work Sheet Cols:", ws.max_column


# 建立存储数据的字典
data_dic = {}

# 把数据存到字典中
for rx in range(1, ws.max_row + 1):
    temp_list = []
    pid = rx
    for cx in range(1,ws.max_column+1):
        w1 = ws.cell(row=rx, column=cx).value
        temp_list.append(w1)
    data_dic[pid] = temp_list

# 打印字典数据个数
print 'Total:%d' % len(data_dic)
print json.dumps(data_dic, encoding="UTF-8", ensure_ascii=False)


new_wb = openpyxl.Workbook()
#new_sheet = new_wb.active
#new_sheet.title = "sheetName"
sheet_1 = new_wb.create_sheet(index=0, title=unicode('主商品', "utf8"))
sheet_1.cell(row=1,column=1).value="aaa"
sheet_2 = new_wb.create_sheet(index=1, title=unicode('子商品', "utf8"))
sheet_3 = new_wb.create_sheet(index=2, title=unicode('子商品规格属性', "utf8"))

new_wb.save("testWrite.xlsx");



