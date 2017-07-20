#!/usr/bin/env python 
# -*- coding: utf-8 -*-

from openpyxl.reader.excel import load_workbook
import json,openpyxl

wb = load_workbook(filename=r'memeqianbao-2017-06-02.xlsx')

# 显示有多少张表
print "Worksheet range(s):", wb.get_named_ranges()
print "Worksheet name(s):", wb.get_sheet_names()

sheetnames = wb.get_sheet_names()
ws = wb.get_sheet_by_name(sheetnames[0])
cell_11 = ws["A1"]
print("cess_11 = "+cell_11.value) 
# 显示表名，表行数，表列数
print "Work Sheet Titile:", ws.title
print "Work Sheet Rows:", ws.max_row
print "Work Sheet Cols:", ws.max_column


# 建立存储数据的字典
dataList = []

# 把数据存到字典中
for rx in range(1, ws.max_row + 1):
    temp_list = []
    for cx in range(1,ws.max_column+1):
        cellValue = ws.cell(row=rx, column=cx).value
        temp_list.append(cellValue)
    dataList.append(temp_list)

# 打印字典数据个数
print 'Total:%d' % len(dataList)
#print json.dumps(dataList, encoding="UTF-8", ensure_ascii=False)

sheet1_name="主商品"
sheet1_data=[]
sheet1_header=["商户id","供货商商品id","商品名称","展示名称","一级品类id","二级品类id","配送范围","配送费","购买说明","图文描述","规格参数","退换货业务介绍","商品图片"]
sheet1_data.append(sheet1_header)


sheet2_name="子商品"
sheet2_data=[]
sheet2_header=["供货商子商品id","供货商商品id","SKU码","SKU名称","价格","库存"]
sheet2_data.append(sheet2_header)

sheet3_name="子商品规格属性"
sheet3_data=[]
sheet3_header=["供货商子商品id","供货商商品id","属性名","属性值"]
sheet3_data.append(sheet3_header)



#for i in range(0,len(dataList)):
for i in range(0,2)):
    print("-----------------------------")
    sheet1_data_temp = []
    sheet2_data_temp = []
    sheet3_data_temp = []
    sheet1_data_temp.insert(2,dataList[i][1]) #商品名称
    sheet1_data_temp.insert(1,dataList[i][3]) #商品id



#new_wb = openpyxl.Workbook()
#new_sheet = new_wb.active
#new_sheet.title = "sheetName"
#sheet_1 = new_wb.create_sheet(index=0, title=unicode('主商品', "utf8"))
#sheet_1.cell(row=1,column=1).value=dataList[0][0]
#sheet_2 = new_wb.create_sheet(index=1, title=unicode('子商品', "utf8"))
#sheet_3 = new_wb.create_sheet(index=2, title=unicode('子商品规格属性', "utf8"))
#new_wb.save("testWrite.xlsx");